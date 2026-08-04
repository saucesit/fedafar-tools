#!/usr/bin/env python3
"""
camara_reporte.py — Genera el PDF del reporte de auditoría de cadena de frío
a partir del resumen (min/max/promedio por sensor) que devuelve la función
SQL reporte_camara. Marca en rojo los sensores de cámara (2-8°C) que se salieron
de rango.
"""

import io, unicodedata
from datetime import datetime
from fpdf import FPDF


def generar_excel(rows):
    """Genera la planilla Excel con el MISMO formato que exporta ESPDesign
    (hoja 'Datos', columnas Fecha/Dispositivo/Sensor/Valor/Unidad), la que se
    presenta a ANMAT. `rows`: lista de dicts con fecha, dispositivo,
    sensor_nombre, valor (ya ordenados como se quieran mostrar). Bytes .xlsx."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Datos'
    ws.append(['Fecha', 'Dispositivo', 'Sensor', 'Valor', 'Unidad'])
    for r in rows:
        val = r.get('valor')
        val_txt = ('%g' % val) if val is not None else ''
        fecha = str(r.get('fecha') or '').replace('T', ' ')[:19]
        ws.append([fecha, r.get('dispositivo') or '', r.get('sensor_nombre') or '',
                   val_txt, ' °C'])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _norm(s):
    t = unicodedata.normalize('NFKD', str(s or ''))
    return ''.join(c for c in t if not unicodedata.combining(c)).lower()


def rango_esperado(sensor_nombre):
    """Rango aceptable °C según el tipo de sensor. Cámara de frío = 2 a 8."""
    n = _norm(sensor_nombre)
    if 'camara' in n:
        return (2.0, 8.0)
    return None   # depósito/ambiente: sin rango estricto por ahora


def _fmt_fecha(f):
    if not f:
        return '-'
    s = str(f).replace('T', ' ')
    return s[:16]


def generar_pdf(desde, hasta, filas):
    """filas: lista de dicts con dispositivo, sensor_nombre, lecturas, minimo,
    maximo, promedio, primera, ultima. Devuelve bytes del PDF."""
    pdf = FPDF(unit='mm', format='A4')
    pdf.set_auto_page_break(True, margin=15)
    pdf.add_page()

    # Encabezado
    pdf.set_font('Helvetica', 'B', 15)
    pdf.set_text_color(0, 74, 153)
    pdf.cell(0, 9, 'Reporte de Cadena de Frio - FEDAFAR', ln=1)
    pdf.set_font('Helvetica', '', 10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 6, f'Periodo: {desde}  a  {hasta}', ln=1)
    pdf.cell(0, 6, f'Generado: {datetime.now():%d/%m/%Y %H:%M}', ln=1)
    pdf.ln(3)

    if not filas:
        pdf.set_font('Helvetica', '', 11)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 8, 'No hay registros de temperatura en el periodo seleccionado.', ln=1)
        return bytes(pdf.output())

    # Tabla
    cols = [('Equipo / Sensor', 62), ('Lecturas', 20), ('Min', 18), ('Max', 18),
            ('Prom', 18), ('Rango OK', 20), ('Estado', 24)]
    pdf.set_font('Helvetica', 'B', 8.5)
    pdf.set_fill_color(0, 74, 153)
    pdf.set_text_color(255, 255, 255)
    for titulo, w in cols:
        pdf.cell(w, 7, titulo, border=1, align='C', fill=True)
    pdf.ln()

    pdf.set_font('Helvetica', '', 8)
    for r in filas:
        rango = rango_esperado(r.get('sensor_nombre'))
        mn, mx = r.get('minimo'), r.get('maximo')
        fuera = False
        if rango and mn is not None and mx is not None:
            fuera = (mn < rango[0]) or (mx > rango[1])
        # Color de fila
        if fuera:
            pdf.set_fill_color(254, 226, 226); estado = 'DESVIO'; ec = (153, 27, 27)
        elif rango:
            pdf.set_fill_color(240, 253, 244); estado = 'OK'; ec = (6, 95, 70)
        else:
            pdf.set_fill_color(249, 250, 251); estado = '-'; ec = (110, 110, 110)

        nombre = f"{r.get('dispositivo','')[:26]} / {r.get('sensor_nombre','')}"
        rango_txt = f"{rango[0]:.0f}-{rango[1]:.0f}C" if rango else '-'
        pdf.set_text_color(30, 30, 30)
        pdf.cell(cols[0][1], 6, nombre[:40], border=1, fill=True)
        pdf.cell(cols[1][1], 6, str(r.get('lecturas', '')), border=1, align='C', fill=True)
        pdf.cell(cols[2][1], 6, f"{mn:.1f}" if mn is not None else '-', border=1, align='C', fill=True)
        pdf.cell(cols[3][1], 6, f"{mx:.1f}" if mx is not None else '-', border=1, align='C', fill=True)
        pdf.cell(cols[4][1], 6, f"{r.get('promedio'):.1f}" if r.get('promedio') is not None else '-', border=1, align='C', fill=True)
        pdf.cell(cols[5][1], 6, rango_txt, border=1, align='C', fill=True)
        pdf.set_text_color(*ec)
        pdf.set_font('Helvetica', 'B', 8)
        pdf.cell(cols[6][1], 6, estado, border=1, align='C', fill=True)
        pdf.set_font('Helvetica', '', 8)
        pdf.ln()

    pdf.ln(4)
    pdf.set_font('Helvetica', 'I', 7.5)
    pdf.set_text_color(120, 120, 120)
    pdf.multi_cell(0, 4, 'Rango de referencia para camara de frio: 2 a 8 C. "DESVIO" indica '
                         'que el minimo o el maximo del periodo se salio de ese rango. Los '
                         'sensores de deposito/ambiente se listan sin rango estricto. Datos '
                         'registrados por los dataloggers ESPDesign, conservados en FEDAFAR.')
    return bytes(pdf.output())
